import time

import openpyxl
import torch
from train_eval import evaluate
from importlib import import_module
import argparse
from utils import build_dataset, build_iterator, get_time_dif
from sklearn.metrics import f1_score

parser = argparse.ArgumentParser(description='Chinese Text Classification')
parser.add_argument('--model', type=str, required=True, help='choose a model: Bert, ERNIE')
parser.add_argument('--test', type=str, required=False, help='choose a testType :single or file')
parser.add_argument('--data', type=str, required=True, help='choose a dataset :single or file')
args = parser.parse_args()

classDic = {0: "财经",
            1: "房产",
            2: "教育",
            3: "科技",
            4: "军事",
            5: "汽车",
            6: "体育",
            7: "游戏",
            8: "娱乐",
            9: "其他",
            }


def singleTest(model, config, text):
    start_time = time.time()
    config.batch_size = 1
    config.text = text
    test_data = build_dataset(config, True)
    test_iter = build_iterator(test_data, config)
    model.eval()
    test_acc, test_loss, test_report, test_confusion, predict, label = evaluate(config, model, test_iter, test=True)
    print(predict[0])
    end_time = time.time()
    print("Time usage:", int(round(end_time * 1000 - start_time * 1000)))


def fileTest(model, config, filepath=""):
    if filepath != "":
        config.test_path = filepath
    test_data = build_dataset(config, True)
    start_time = time.time()
    test_iter = build_iterator(test_data, config)
    model.eval()
    test_acc, test_loss, test_report, test_confusion, predict, label = evaluate(config, model, test_iter, test=True)
    file = open(dataset + "/output/" + model_name + "OutPut.txt", "w+", encoding='utf-8')
    file1 = open(dataset + "/data/test.txt", encoding='utf-8')
    for line, pre in zip(file1, predict):
        file.write(line.split('\t@')[0] + '@' + classDic[int(line.split('\t@')[1][0])] + ',' + classDic[pre] + '\n')
    msg = 'Test Loss: {0:>5.2},  Test Acc: {1:>6.2%}'
    print(msg.format(test_loss, test_acc))
    print("Precision, Recall and F1-Score...")
    print(test_report)
    print("Confusion Matrix...")
    print(test_confusion)
    f1_micro = f1_score(label, predict, average='micro')
    print("F1_Score:")
    print(f1_micro)
    time_dif = get_time_dif(start_time)
    print("Time usage:", time_dif)


def fileTest1(model, config, filepath=""):
    filepath = "test.xlsx"
    test_data = build_dataset(config, True)
    start_time = time.time()
    test_iter = build_iterator(test_data, config)
    model.eval()
    test_acc, test_loss, test_report, test_confusion, predict, label = evaluate(config, model, test_iter, test=True)
    if filepath.split(".")[1] == "xlsx":
        workBook = openpyxl.load_workbook(filepath)
        sheets = workBook.sheetnames
        sheet = workBook[sheets[0]]
        l = 0
        for row, pre in zip(sheet.iter_rows(), predict):
            l += 1
            row[1].value = classDic[pre]
        workBook.save(filepath)
        workBook.close()
    end_time = time.time()
    print("已完成!!!\n总用时(秒):" + str(float(round(end_time * 1000 - start_time * 1000)) / 1000))
    print("Time usage:", int(round(end_time * 1000 - start_time * 1000)))
    print("finished")


if __name__ == '__main__':
    dataset = args.data  # 数据集
    model_name = args.model  # bert
    x = import_module('models.' + model_name)
    config = x.Config(dataset)
    model = x.Model(config).to(config.device)
    model.load_state_dict(torch.load(config.save_path, map_location='cuda' if torch.cuda.is_available() else 'cpu'))
    if args.test == 'text':
        singleTest(model, config, "《高殿战记》再次迎来一波小更新。此次更新内容丰富，制作组诚意拉满。让我们来一睹为快！开发者收到多位玩家反馈，更新了制作界面。物品等	@7")
    else:
        fileTest1(model, config, config.test_path)
